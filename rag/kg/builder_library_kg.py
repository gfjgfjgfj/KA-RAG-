from rag.kg.base_build import Builder
from openpyxl import load_workbook


class LibraryKGBuilder(Builder):
    def __init__(self):
        super().__init__("bolt://localhost:7687", "neo4j", "08102775")

    def test(self):
        # result = self.g.run("MATCH (n:course)-[r]-(o) WHERE n.name = '大学英语A1' RETURN n.name, n.course_url, n.course_type, n.course_code, r.name,o.name")
        # for item in result:
        #     print(result)
        print(self.getCourseDetail("大学英语A1"))
    def build(self):
        # result = self.g.run("MATCH (p:course {name: '大学英语A1'})-[r]->(other) RETURN p, r, other")
        # print(result)
        workbook = load_workbook('all_courses.xlsx')
        sheet = workbook['Sheet1']
        self.addCourseMajorRelationship(sheet)

    def buildCourse(self, sheet):
        for index, row in enumerate(sheet.iter_rows(min_row=2, max_col=1)):
            url = []
            for col in sheet.iter_cols(min_col=3, max_col=7, max_row=index, min_row=index):
                value = col[0].value
                if value:
                    url.append(value)
            self.create_node(label="course", node=row[0].value, course_url=url)

    def buildDepartment(self, sheet):
        ls = []
        for index, row in enumerate(sheet.iter_rows(min_row=3, max_col=1)):
            ls.append(row[0].value)
        for index, row in enumerate(sheet.iter_rows(min_row=3, min_col=6, max_col=6)):
            ls.append(row[0].value)
        for obj in set(ls):
            if obj:
                self.create_node(label="department", node=obj)

    def buildMajor(self, sheet):
        ls = []
        last = None
        for row in sheet.iter_rows(min_row=3, min_col=2, max_col=2):
            obj = row[0].value
            if obj:
                if last != obj:
                    ls.append(obj)
                    last = obj
        self.create_nodes(label="major", nodes=ls)

    def addCourseProperties(self, sheet):
        for index, row in enumerate(sheet.iter_rows(min_row=3, min_col=4, max_col=7)):
            if row[0].value:
                self.addProperties(label="course", node=row[1].value, course_code=row[0].value,
                                   course_type=row[3].value)

    def addCourseStudentRelationship(self, sheet):
        set_edges = []
        for row in sheet.iter_rows(min_row=3, min_col=1, max_col=6):
            if row[0].value:
                student_department = row[0].value
                # teacher_department = row[5].value
                course = row[4].value
                set_edges.append(f"{student_department}-{course}")
                # related_majors = row[1].value

        for it in set(set_edges):
            ls = it.split("-")
            self.create_relationship(
                start_label="course",
                end_label="department",
                start_node=ls[1],
                end_node=ls[0],
                rel_type="student_department",
                rel_name="上课学院"
            )
    def addCourseTeacherRelationship(self, sheet):
        set_edges = []
        for row in sheet.iter_rows(min_row=3, min_col=1, max_col=6):
            if row[0].value:
                teacher_department = row[5].value
                course = row[4].value
                set_edges.append(f"{teacher_department}-{course}")
        for it in set(set_edges):
            ls = it.split("-")
            self.create_relationship(
                start_label="course",
                end_label="department",
                start_node=ls[1],
                end_node=ls[0],
                rel_type="teacher_department",
                rel_name="开课学院"
            )

    def addCourseMajorRelationship(self, sheet):
        set_edges = []
        for row in sheet.iter_rows(min_row=3, min_col=1, max_col=6):
            if row[0].value:
                related_majors = row[1].value
                course = row[4].value
                set_edges.append(f"{related_majors}-{course}")
        for it in set(set_edges):
            ls = it.split("-")
            self.create_relationship(
                start_label="course",
                end_label="major",
                start_node=ls[1],
                end_node=ls[0],
                rel_type="student_major",
                rel_name="上课学生的专业"
            )

    def getUrlByCourse(self, course_name):
        result = self.g.run(
            f"MATCH (n:course) WHERE n.name = '{course_name}' RETURN n.course_url")
        data = result.data()
        if data is not None:
            if len(data) != 0:
                return data[0]['n.course_url']
        return ""

    def getTeacherDepartmentByCourse(self,course_name):
        result = self.g.run(
            f"MATCH (n:course)-[r]-(o) WHERE n.name = '{course_name}' and r.name = '开课学院' RETURN o.name")
        data = result.data()
        if data is not None:
            if len(data) != 0:
                return data[0]['o.name']
        return ""

    def getStudentDepartmentByCourse(self, course_name):
        result = self.g.run(
            f"MATCH (n:course)-[r]-(o) WHERE n.name = '{course_name}' and r.name = '上课学院' RETURN o.name")
        data = result.data()
        re = []
        if data is not None:
            if len(data) != 0:
                for item in data:
                    re.append(item['o.name'])
        return re

    def getStudentMajorByCourse(self, course_name):
        result = self.g.run(
            f"MATCH (n:course)-[r]-(o) WHERE n.name = '{course_name}' and r.name = '上课学生的专业' RETURN o.name")
        data = result.data()
        re = []
        if data is not None:
            if len(data) != 0:
                for item in data:
                    re.append(item['o.name'])
        return re

    def getCourseDetail(self, course_name):
        course_info = self.g.run(f"MATCH (n:course) WHERE n.name=~'(?i).*{course_name}.*' RETURN n.course_url, n.course_code,n.name").data()
        if len(course_info) != 0:
            item = course_info[0]
            urls = item["n.course_url"]
            course_code = item["n.course_code"]
            result_course_name = item["n.name"]
        else:
            return None
        major_info = self.g.run(
            f"MATCH (n:course)-[r]-(o) WHERE n.name ='{result_course_name}' and r.name = '上课学生的专业' RETURN o limit 3").data()
        majors = []
        for item in major_info:
            for value in item["o"].values():
                majors.append(value)
        teachers = []
        teacher_info = self.g.run(
            f"MATCH (n:course)-[r]-(o) WHERE n.name = '{result_course_name}' and r.name = '开课学院' RETURN o").data()
        for item in teacher_info:
            for value in item["o"].values():
                teachers.append(value)
        students = []
        student_info = self.g.run(
            f"MATCH (n:course)-[r]-(o) WHERE n.name = '{result_course_name}' and r.name = '上课学院' RETURN o limit 3").data()
        for item in student_info:
            for value in item["o"].values():
                students.append(value)
        first = f"{result_course_name}课程信息：课程编码{course_code}，由{"、".join(teachers)}开设，课程面向{"、".join(students)}等学院的学生，专业包括{"、".join(majors)}等,课程在线学习链接：{urls}。"
        second = self.get_knowledge(course_name=result_course_name)
        if second:
            return first + "\n" + second
        else:
            return first
    def get_knowledge(self, course_name):
        result = []
        info = self.g.run(f"MATCH (n:course)-[r:content]->(o) WHERE n.name = '{course_name}' RETURN n.chapter_num, "
                          f"r.name, o.name").data()
        if len(info) == 0:
            return None
        else:
            result.append(f"{course_name}知识点：")
            for it in info:
                result.append(f"{it["r.name"]}：${it["o.name"]}")
                knowledge_info = self.g.run(
                    f"MATCH (n:chapter)-[r:content]->(o) WHERE n.name = '{it["o.name"]}' RETURN o").data()
                if len(knowledge_info) == 0:
                    return None
                ls = []
                for it2 in knowledge_info:
                    temp = it2["o"]["name"]# + "：介绍：" + it2["o"]["introduction"]
                    # if it2["o"]["application"]:
                    #     temp = temp + " 应用场景：" + it2["o"]["application"]
                    # if it2["o"]["reference_link"]:
                    #     temp = temp + " 参考链接：" + it2["o"]["reference_link"]
                    #if it2["o"]["video_link"]:
                    #     temp = temp + " 视频学习：" + it2["o"]["video_link"]
                    # if it2["o"]["paper"]:
                    #     temp = temp + " 相关论文：" + it2["o"]["paper"]
                    ls.append(temp)
                result.append("包含知识点：" + "、".join(ls))
        return "\n".join(result)

    def get_knowledge_detail(self, knowledge):
        result = []
        info = self.g.run(f"MATCH (o:knowledge) WHERE o.name='{knowledge}' RETURN o").data()
        if len(info) == 0:
            return None
        else:
            for it in info:
                temp = it["o"]["name"] + "：介绍：" + it["o"]["introduction"]
                if it["o"]["application"]:
                    temp = temp + " 应用场景：" + it["o"]["application"]
                if it["o"]["reference_link"]:
                    temp = temp + " 参考链接：" + it["o"]["reference_link"]
                if it["o"]["video_link"]:
                    temp = temp + " 视频学习：" + it["o"]["video_link"]
                if it["o"]["paper"]:
                    temp = temp + " 相关论文：" + it["o"]["paper"]
                result.append(temp)
        return "\n".join(result)